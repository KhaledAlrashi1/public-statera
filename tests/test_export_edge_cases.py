import csv
import io
import unittest
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from backend.constants import EXPORT_CSV_MAX_ROWS
from preflight_base import PreflightApiTestBase


class ExportEdgeCaseTests(PreflightApiTestBase):
    def _seed_transactions(self, *, user_id: int, count: int) -> None:
        with self.app.app_context():
            category = self.Category.query.filter_by(user_id=user_id, name="Export Edge").first()
            if category is None:
                category = self.Category(user_id=user_id, name="Export Edge", is_income=False)
                self.db.session.add(category)
                self.db.session.flush()

            batch_size = 1000
            for start in range(0, count, batch_size):
                txns = []
                for idx in range(start, min(start + batch_size, count)):
                    txns.append(
                        self.Transaction(
                            user_id=user_id,
                            date=date(2026, 2, 1),
                            category_id=category.id,
                            name=f"Edge Txn {idx}",
                            name_key=f"edge-txn-{idx}",
                            amount_kd=Decimal("1.000"),
                        )
                    )
                self.db.session.add_all(txns)
                self.db.session.flush()
            self.db.session.commit()

    def _csv_rows(self, response) -> list[list[str]]:
        content = response.get_data(as_text=True)
        if content.startswith("\ufeff"):
            content = content[1:]
        return list(csv.reader(io.StringIO(content)))

    def test_export_csv_zero_transactions_returns_header_only(self):
        self._create_user("export-empty@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-empty@example.com", "Password123!")

        res = client.get("/api/transactions/export-csv")
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        rows = self._csv_rows(res)
        self.assertEqual(len(rows), 1)
        self.assertEqual(
            rows[0],
            ["transaction_id", "date", "merchant", "category", "name", "amount_kd", "memo"],
        )
        self.assertEqual(res.headers.get("X-Export-Truncated"), "false")

    def test_export_csv_exact_row_limit_reports_truncated_true(self):
        user_id = self._create_user("export-exact-limit@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-exact-limit@example.com", "Password123!")
        self._seed_transactions(user_id=user_id, count=EXPORT_CSV_MAX_ROWS)

        res = client.get("/api/transactions/export-csv")
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        rows = self._csv_rows(res)
        self.assertEqual(len(rows), EXPORT_CSV_MAX_ROWS + 1)  # header + data
        self.assertEqual(res.headers.get("X-Export-Truncated"), "true")
        self.assertEqual(res.headers.get("X-Export-Row-Limit"), str(EXPORT_CSV_MAX_ROWS))

    def test_export_csv_uses_transaction_amounts_for_atomic_rows(self):
        user_id = self._create_user("export-grouped@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-grouped@example.com", "Password123!")

        with self.app.app_context():
            category = self.Category(user_id=user_id, name="Groceries", is_income=False)
            merchant = self.Merchant(user_id=user_id, name="Local Market")
            self.db.session.add_all([category, merchant])
            self.db.session.flush()

            txn = self.Transaction(
                user_id=user_id,
                date=date(2026, 2, 3),
                category_id=category.id,
                merchant_id=merchant.id,
                name="Weekly Shop",
                memo="Imported legacy split",
                name_key="weekly-shop",
                amount_kd=Decimal("4.000"),
            )
            self.db.session.add(txn)
            self.db.session.commit()

        res = client.get("/api/transactions/export-csv")
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        rows = self._csv_rows(res)
        self.assertEqual(
            rows,
            [
                ["transaction_id", "date", "merchant", "category", "name", "amount_kd", "memo"],
                ["1", "2026-02-03", "Local Market", "Groceries", "Weekly Shop", "4.000", "Imported legacy split"],
            ],
        )

    def test_export_xlsx_uses_canonical_flat_schema(self):
        user_id = self._create_user("export-xlsx@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-xlsx@example.com", "Password123!")
        self._seed_transactions(user_id=user_id, count=1)

        res = client.get("/api/transactions/export-xlsx")
        self.assertEqual(res.status_code, 200, res.get_data(as_text=False))
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            res.content_type,
        )

        workbook = load_workbook(io.BytesIO(res.data), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(
            list(rows[0]),
            ["transaction_id", "date", "merchant", "category", "name", "amount_kd", "memo"],
        )
        self.assertEqual(rows[1][1:], ("2026-02-01", None, "Export Edge", "Edge Txn 0", "1.000", None))
        self.assertTrue(rows[1][0])

    def test_export_csv_sanitizes_formula_like_cells(self):
        user_id = self._create_user("export-sanitize@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-sanitize@example.com", "Password123!")

        with self.app.app_context():
            category = self.Category(user_id=user_id, name="=Formula Category", is_income=False)
            merchant = self.Merchant(user_id=user_id, name="+Formula Merchant")
            self.db.session.add_all([category, merchant])
            self.db.session.flush()

            txn = self.Transaction(
                user_id=user_id,
                date=date(2026, 2, 7),
                category_id=category.id,
                merchant_id=merchant.id,
                name="@Formula Name",
                memo="-Formula Memo",
                name_key="formula-export",
                amount_kd=Decimal("3.500"),
            )
            self.db.session.add(txn)
            self.db.session.commit()

        res = client.get("/api/transactions/export-csv")
        self.assertEqual(res.status_code, 200, res.get_data(as_text=True))
        rows = self._csv_rows(res)
        self.assertEqual(
            rows[1],
            ["1", "2026-02-07", "'+Formula Merchant", "'=Formula Category", "'@Formula Name", "3.500", "'-Formula Memo"],
        )

    def test_export_csv_response_is_streamed(self):
        user_id = self._create_user("export-stream-csv@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-stream-csv@example.com", "Password123!")
        self._seed_transactions(user_id=user_id, count=2)

        res = client.get("/api/transactions/export-csv", buffered=False)
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.is_streamed)
        body = "".join(
            chunk.decode("utf-8") if isinstance(chunk, bytes) else chunk
            for chunk in res.response
        )
        self.assertIn("transaction_id,date,merchant,category,name,amount_kd,memo", body)

    def test_export_xlsx_response_is_streamed(self):
        user_id = self._create_user("export-stream-xlsx@example.com", "Password123!")
        client = self.app.test_client()
        self._login(client, "export-stream-xlsx@example.com", "Password123!")
        self._seed_transactions(user_id=user_id, count=2)

        res = client.get("/api/transactions/export-xlsx", buffered=False)
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.is_streamed)
        payload = b"".join(res.response)

        workbook = load_workbook(io.BytesIO(payload), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        self.assertEqual(
            list(rows[0]),
            ["transaction_id", "date", "merchant", "category", "name", "amount_kd", "memo"],
        )


if __name__ == "__main__":
    unittest.main()
